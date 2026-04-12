"""
Excel Data Filter - VLOOKUP-like functionality
Filters students with "Activo" status from the tradicional sheet
pip install pandas openpyxl customtkinter 
python -m PyInstaller --onefile --windowed NominaCreator.py  --hidden-import=pandas --hidden-import=openpyxl --hidden-import=customtkinter --hidden_import=pywin32

"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText

# File configuration
EXCEL_FILE = "1-Registro-2026 TST.xlsm"
SHEET_NAME = "Tradicional"

# Sheet Limits
SHEET_LIMIT = 160
DATA_RANGE = f"B18:AP{SHEET_LIMIT}"
STATUS_COLUMN = "Status"
FILTER_VALUE = "Activo"

# Payroll constants
PR_INDV = 120  # Payment rate for individual classes (per hour)
PR_COM = 160  # Base payment rate for group classes (per session)
PR_ESTR = 60
PR_ASM = 200
# Payroll parameters based on execution date
_today = datetime.now()
PAYROLL_MONTH_NUM = _today.month  # Month number (1-12)
PAYROLL_YEAR = _today.year
_MONTHS_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]
PAYROLL_MONTH = _MONTHS_ES[PAYROLL_MONTH_NUM - 1]
CURRENCY_SYMBOL = "$"
CURRENCY_NAME = "PESOS MXN"

# Professor to Instrument mapping
PROFESSOR_INSTRUMENTS = {
    "Valentin Ramirez de Luna": "Maestro de Canto",
#    "Pavel Boudniak Pugachov": "Maestro de Canto",
    "Oscar Alan Toledo Chavez": "Maestro de Bateria",
    "Estephanie Gallegos Jauregui": "Maestra de Iniciación Musical",
    "Fernando Olmos Camacho": "Maestro de Piano",
    "Joel Alejandro Camacho": "Maestro de Guitarra",
    "Saira Patricia de la Torre": "Maestra de Canto"
}

# Short name to Full name mapping (from Excel to full names)
PROFESSOR_NAME_MAP = {
    "alex": "Joel Alejandro Camacho",
    "oscar": "Oscar Alan Toledo Chavez",
    "valentin": "Valentin Ramirez de Luna",
    "pavel": "Pavel Boudniak Pugachov",
    "estephanie": "Estephanie Gallegos Jauregui",
    "fernando": "Fernando Olmos Camacho",
    "saira": "Saira Patricia de la Torre"
    # Add more mappings as needed
}

def normalize_professor_name(name):
    """Convert shortened professor names to full names using the mapping."""
    if pd.isna(name):
        return name
    name_str = str(name).strip().lower()
    # Check if it matches a key in the name map
    if name_str in PROFESSOR_NAME_MAP:
        return PROFESSOR_NAME_MAP[name_str]
    # Return original if no mapping found
    return str(name).strip()


def load_and_filter_data(file_path, sheet_name, use_range=None):
    """
    Load Excel data and filter by active status.
    """
    try:
        if use_range:
            # Row 17 is header; rows 18-156 contain data when using B18:AP300
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                skiprows=16,
                nrows=SHEET_LIMIT,
            )
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)

        print(f"✓ Loaded {len(df)} total rows from '{sheet_name}' sheet")
        print(f"\nColumns found: {list(df.columns)}\n")

        if STATUS_COLUMN in df.columns:
            filtered_df = df[df[STATUS_COLUMN] == FILTER_VALUE].copy()
            print(f"✓ Filtered {len(filtered_df)} rows with status '{FILTER_VALUE}'")
        else:
            print(f"⚠ Warning: '{STATUS_COLUMN}' column not found; skipping status filter")
            filtered_df = df

        return filtered_df
    except Exception as e:
        print(f"✗ Error loading data: {e}")
        return None


def group_by_professor(df):
    """Group students by professor column (robust to header variants)."""
    if df is None or df.empty:
        return {}

    prof_col = None
    for c in df.columns:
        if isinstance(c, str) and 'PROF' in c.upper():
            prof_col = c
            break
    if prof_col is None and 'PROFESOR' in df.columns:
        prof_col = 'PROFESOR'
    if prof_col is None:
        print("⚠ Warning: could not find a professor column (searched for 'PROF')")
        return {}

    df_clean = df[df[prof_col].notna()].copy()

    groups = {}
    for prof in df_clean[prof_col].unique():
        # Normalize the professor name (convert short names to full names)
        normalized_prof = normalize_professor_name(prof)
        rows = df_clean[df_clean[prof_col] == prof]
        groups[normalized_prof] = rows.to_dict('records')

    return groups
def display_professor_groups(professor_groups, columns_to_show=None):
    """
    Display students grouped by professor
    
    Parameters:
    - professor_groups: Dictionary of professor -> students mapping
    - columns_to_show: List of columns to display for each student
    """
    if not professor_groups:
        print("No professor groups to display")
        return
    
    print("\n" + "="*80)
    print(f"STUDENTS GROUPED BY PROFESSOR")
    print("="*80 + "\n")
    
    for professor, students in sorted(professor_groups.items()):
        print(f"\n{'='*80}")
        print(f" PROFESOR: {professor}")
        print(f"   Total students: {len(students)}")
        print(f"{'='*80}\n")
        
        # Display each student
        for i, student in enumerate(students, 1):
            print(f"  Student {i}:")
            
            if columns_to_show:
                for col in columns_to_show:
                    if col in student and col != "PROFESOR":
                        value = student[col]
                        print(f"    - {col}: {value}")
            else:
                for key, value in student.items():
                    if key != "PROFESOR":
                        print(f"    - {key}: {value}")
            print()
    
    print("="*80 + "\n")


def export_professors_to_separate_files(professor_groups, output_dir="."):
    """
    Export each professor's students to a separate Excel file
    
    Parameters:
    - professor_groups: Dictionary of professor -> students mapping
    - output_dir: Directory to save the files
    """
    if not professor_groups:
        print("No data to export")
        return
    
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    for professor, students in professor_groups.items():
        # Create safe filename
        safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in professor)
        filename = f"{safe_name}_students.xlsx"
        filepath = output_path / filename
        
        # Convert list of dicts back to DataFrame
        df_professor = pd.DataFrame(students)
        
        try:
            df_professor.to_excel(filepath, index=False, sheet_name=professor[:31])  # Excel sheet name limit
            print(f"✓ Exported {len(students)} students for {professor} to '{filename}'")
        except Exception as e:
            print(f"✗ Error exporting {professor}: {e}")
    
    print(f"\n✓ All files saved to: {output_path.absolute()}")


def _safe_time_value(val):
    """Try to normalize a time-like value for sorting and display as HH:MM.

    Handles pandas Timestamps, datetime/time objects and strings like
    'Martes 16:00' or '13:45:00' and returns '13:45'.
    """
    if pd.isna(val):
        return ""

    # Handle pandas Timestamp, datetime or time objects
    try:
        import datetime
        if isinstance(val, pd.Timestamp) or isinstance(val, datetime.datetime):
            t = val.time()
            return t.strftime("%H:%M")
        if isinstance(val, datetime.time):
            return val.strftime("%H:%M")
    except Exception:
        pass

    s = str(val).strip()
    # If contains a space with day like 'Martes 16:00', take the last token
    if " " in s:
        parts = s.split()
        candidate = parts[-1]
    else:
        candidate = s

    # Normalize time strings: remove seconds if present, keep HH:MM
    if ":" in candidate:
        parts = candidate.split(":")
        if len(parts) >= 2:
            hh = parts[0].zfill(2)
            mm = parts[1].zfill(2)
            return f"{hh}:{mm}"

    # fallback: return original string
    return s


def _extract_duration_minutes(row):
    """Extract duration in minutes from row, handling H:MM or numeric values."""
    duration_val = None
    for k in row.keys():
        if k is None:
            continue
        try:
            key_upper = str(k).upper()
        except Exception:
            key_upper = ''
        if 'DUR' in key_upper or 'DURATION' in key_upper or 'MIN' in key_upper:
            v = row.get(k)
            if v not in (None, '', '(en blanco)'):
                duration_val = v
                break
    
    if duration_val is not None:
        dur_minutes = None
        try:
            s = str(duration_val).strip()
            if ':' in s:
                parts = s.split(':')
                hours = float(parts[0]) if parts[0] else 0.0
                mins = float(parts[1]) if len(parts) > 1 and parts[1] else 0.0
                dur_minutes = hours * 60 + mins
            else:
                dur_minutes = float(s.replace(',', '.'))
        except Exception:
            dur_minutes = None
        return dur_minutes
    return None


def _extract_duration_minutes(row):
    """Find a duration-like value in the row and return minutes as float, or None."""
    duration_val = None
    for k in row.keys():
        if k is None:
            continue
        try:
            key_upper = str(k).upper()
        except Exception:
            key_upper = ""
        if "DUR" in key_upper or "DURATION" in key_upper or "MIN" in key_upper:
            v = row.get(k)
            if v not in (None, "", "(en blanco)"):
                duration_val = v
                break

    if duration_val is None:
        return None

    try:
        s = str(duration_val).strip()
        if ":" in s:
            parts = s.split(":")
            hours = float(parts[0]) if parts[0] else 0.0
            mins = float(parts[1]) if len(parts) > 1 and parts[1] else 0.0
            return hours * 60 + mins
        return float(s.replace(",", "."))
    except Exception:
        return None


def generate_payroll_for_professors(professor_groups):
    """
    Build per-professor nomina rows ordered by day/time.

    Rules:
    - No. increments per student per professor.
    - G increments when DÍA/HORARIO change.
    - Clases defaults to 4 when column exists.
    - Estrellita classes: Pago Prop = 60.
    - Ensamble Pago Prop = 50;
    - Comp classes (size >= 2): Pago Prop computed once per (DÍA, HORARIO) group
      as 160 + 40 * (size - 2); other students in the group get 0.
    - Comp classes (size == 1): treat as 30-minute individual (Pago Prop = 60).
    - Individual classes: Pago Prop = 120 * (Duración / 60) when duration found.
    """
    result = {}
    for prof, students in professor_groups.items():
        if not students:
            result[prof] = []
            continue

        df_prof = pd.DataFrame(students)

        # Determine sort columns (prefer DÍA and HORARIO)
        sort_cols = []
        if "DÍA" in df_prof.columns and "HORARIO" in df_prof.columns:
            sort_cols = ["DÍA", "HORARIO"]
        else:
            for c in ("Date", "DÍA", "HORARIO"):
                if c in df_prof.columns:
                    sort_cols.append(c)

        df_prof["_sort_horario"] = df_prof.get("HORARIO", "").apply(_safe_time_value)

        try:
            if sort_cols:
                sort_by = [c for c in sort_cols if c != "HORARIO"] + ["_sort_horario"]
                df_prof = df_prof.sort_values(by=sort_by, na_position="last")
            else:
                df_prof = df_prof.sort_values(by=["_sort_horario"], na_position="last")
        except Exception:
            pass

        nomina = [
            "No.",
            "ALUMNO",
            "DÍA",
            "HORARIO",
            "CLASE",
            "DURACIÓN",
            "Clases",
            "Pago Prop",
            "Monto",
            "RECIBO ACTUAL",
        ]

        chosen_cols = []
        cols_lower = {c: c.lower() for c in df_prof.columns if isinstance(c, str)}
        for p in nomina:
            if p in df_prof.columns:
                chosen_cols.append(p)
                continue
            lower_p = p.lower()
            match = next((orig for orig, low in cols_lower.items() if low == lower_p), None)
            if match:
                chosen_cols.append(match)
                continue
            if p in ("No.", "Clases", "Pago Prop", "Monto"):
                chosen_cols.append(p)

        seen = set()
        chosen_cols = [c for c in chosen_cols if not (c in seen or seen.add(c))]

        if "No." not in chosen_cols:
            chosen_cols.insert(0, "No.")
        if "G" not in chosen_cols:
            chosen_cols.insert(1 if "No." in chosen_cols else 0, "G")

        # Precompute group sizes per (DÍA, HORARIO)
        group_keys = []
        for _, r in df_prof.iterrows():
            d = str(r.get("DÍA", "")).strip() if pd.notna(r.get("DÍA", None)) else ""
            h_raw = r.get("HORARIO", r.get("Horario", ""))
            h = _safe_time_value(h_raw) if h_raw not in (None, "") else ""
            group_keys.append((d, h))

        group_counts = {}
        for k in group_keys:
            group_counts[k] = group_counts.get(k, 0) + 1

        nomina_rows = []
        prev_dia = None
        prev_horario = None
        group_index = 0
        student_no = 0
        assigned_comp_groups = set()

        for _, r in df_prof.iterrows():
            student_no += 1
            row = {c: (r[c] if c in r.index else None) for c in chosen_cols}
            row["No."] = student_no

            # Normalize any horario/time fields to HH:MM string
            try:
                if "HORARIO" in row and row.get("HORARIO", None) not in (None, ""):
                    row["HORARIO"] = _safe_time_value(row["HORARIO"])
                elif "Horario" in row and row.get("Horario", None) not in (None, ""):
                    row["Horario"] = _safe_time_value(row["Horario"])
            except Exception:
                pass

            if "Pago Prop" in row and row["Pago Prop"] in (None, "", "(en blanco)"):
                row["Pago Prop"] = 0
            if "Clases" in row:
    
                row["Clases"] = 4

            cur_dia = str(row.get("DÍA", "")).strip() if row.get("DÍA", None) not in (None, "") else ""
            cur_hor_raw = row.get("HORARIO", row.get("Horario", ""))
            cur_hor = _safe_time_value(cur_hor_raw) if cur_hor_raw not in (None, "") else ""

            if cur_dia != prev_dia or cur_hor != prev_horario:
                group_index += 1
            row["G"] = group_index

            class_val = None
            for key in ("CLASE", "Clase", "TIPO DE CLASE"):
                if key in row and row[key] not in (None, ""):
                    class_val = str(row[key]).strip()
                    break

            handled = False
            if class_val:
                lc = class_val.lower()
                if "estrellita" in lc:
                    row["Pago Prop"] = PR_ESTR
                    handled = True
                elif "ensamble" in lc:
                    k = (cur_dia, cur_hor)
                    size = group_counts.get(k, 1)
                    if size >= 2:
                        pago_value = PR_ASM
                        if k not in assigned_comp_groups:
                            row["Pago Prop"] = pago_value
                            assigned_comp_groups.add(k)
                            row["Clases"] = 2
                        else:
                            row["Pago Prop"] = 0
                        handled = True


                elif "comp" in lc:
                    k = (cur_dia, cur_hor)
                    size = group_counts.get(k, 1)
                    if size >= 2:
                        pago_value = PR_COM + max(0, size - 2) * 40
                        if k not in assigned_comp_groups:
                            row["Pago Prop"] = pago_value
                            assigned_comp_groups.add(k)
                        else:
                            row["Pago Prop"] = 0
                    else:
                        row["Pago Prop"] = 120 * (30.0 / 60.0)
                    handled = True

            if not handled:
                dur_minutes = _extract_duration_minutes(row)
                if dur_minutes is not None:
                    try:
                        row["Pago Prop"] = PR_INDV * (dur_minutes / 60.0)
                    except Exception:
                        pass

            prev_dia = cur_dia
            prev_horario = cur_hor
            nomina_rows.append(row)

        result[prof] = nomina_rows

    return result


def export_nomina_workbooks(professor_nominas, output_dir="nominas"):
    """
    Export each professor's nomina to a separate Excel workbook.
    Each workbook will contain a single sheet named 'Nomina'.
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    for prof, rows in professor_nominas.items():
        if not rows:
            continue
        df = pd.DataFrame(rows)
        safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in prof)
        filename = out / f"nomina_{safe_name}.xlsx"
        try:
            df.to_excel(filename, index=False, sheet_name='Nomina')
            print(f"✓ Exported nómina for {prof} -> {filename}")
        except Exception as e:
            print(f"✗ Error exporting nómina for {prof}: {e}")


def export_xlsx_to_pdf(xlsx_path, pdf_path=None):
    """Export an Excel workbook to PDF using Excel COM on Windows.

    Returns: (success: bool, message_or_output_path: str)
    """
    xlsx_path = Path(xlsx_path)
    if pdf_path is None:
        pdf_path = xlsx_path.with_suffix(".pdf")
    pdf_path = Path(pdf_path)

    try:
        import importlib
        win32_client = importlib.import_module("win32com.client")
    except Exception:
        return False, "pywin32 is not installed. Install with: pip install pywin32"

    excel = None
    wb = None
    try:
        excel = win32_client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(xlsx_path.resolve()))
        # 0 => xlTypePDF, respect each sheet print area/setup.
        wb.ExportAsFixedFormat(0, str(pdf_path.resolve()))
        wb.Close(SaveChanges=False)
        excel.Quit()
        return True, str(pdf_path)
    except Exception as e:
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        return False, f"Error exporting PDF: {e}"


def export_nomina_single_workbook(professor_nominas, output_file="nomina_por_profesor.xlsx"):
    """
    Export all professors' nóminas into a single Excel workbook, one sheet per professor.
    Format: headers at row 5, TOTAL at row 6, data from row 7 onwards, footer info at bottom.
    """
    if not professor_nominas:
        print("No nominas to export")
        return

    out_path = Path(output_file)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.worksheet.page import PageMargins
        from datetime import datetime, timedelta
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # First sheet: summary in requested layout (D:G)
        resumen = wb.create_sheet(title="Resumen", index=0)
        thin = Side(style="thin", color="000000")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header area
        resumen.merge_cells('D2:G2')
        resumen['D2'] = f"Nómina de Maestros {PAYROLL_YEAR}"
        resumen['D2'].font = Font(bold=True, size=12)
        resumen['D2'].border = thin_border
        resumen['E2'].border = thin_border
        resumen['F2'].border = thin_border
        resumen['G2'].border = thin_border
        
        resumen['D2'].alignment = Alignment(horizontal='center')

        resumen.merge_cells('D3:G3')
        resumen['D3'] = PAYROLL_MONTH
        resumen['D3'].font = Font(bold=True, size=11)
        resumen['D3'].border = thin_border
        resumen['E3'].border = thin_border
        resumen['F3'].border = thin_border
        resumen['G3'].border = thin_border
        
        resumen['D3'].alignment = Alignment(horizontal='center')

        # G4: date as 05-{month+1}-{yy}
        next_month = PAYROLL_MONTH_NUM + 1 if PAYROLL_MONTH_NUM < 12 else 1
        year_2d = PAYROLL_YEAR % 100
        resumen['G4'] = f"05/{next_month:02d}/{year_2d:02d}"
        resumen['G4'].font = Font(bold=True)
        resumen['G4'].alignment = Alignment(horizontal='center')

        # Table header at D6:G6
        tbl_headers = ["Maestro", "Pago", "Alumnos", "Recibido"]
        for idx, h in enumerate(tbl_headers, start=4):
            cell = resumen.cell(row=6, column=idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        def _to_float_safe(v):
            try:
                if v in (None, "", "(en blanco)"):
                    return 0.0
                return float(v)
            except Exception:
                return 0.0

        summary_row = 7
        total_payroll = 0.0
        total_students = 0

        for prof_name, prof_rows in sorted(professor_nominas.items()):
            students_count = len(prof_rows)
            payroll_total = 0.0

            for rr in prof_rows:
                clases = _to_float_safe(rr.get("Clases", 0))
                pago_prop = _to_float_safe(rr.get("Pago Prop", 0))
                payroll_total += clases * pago_prop

            resumen.cell(row=summary_row, column=4, value=prof_name)
            resumen.cell(row=summary_row, column=5, value=payroll_total)
            resumen.cell(row=summary_row, column=6, value=students_count)
            resumen.cell(row=summary_row, column=7, value="")
            resumen.cell(row=summary_row, column=5).number_format = '"$" #,##0.0'
            for col in range(4, 8):
                resumen.cell(row=summary_row, column=col).alignment = Alignment(horizontal='center')
                resumen.cell(row=summary_row, column=col).border = thin_border

            total_payroll += payroll_total
            total_students += students_count
            summary_row += 1

        resumen.cell(row=summary_row, column=4, value="∑")
        resumen.cell(row=summary_row, column=5, value=total_payroll)
        resumen.cell(row=summary_row, column=6, value=total_students)
        resumen.cell(row=summary_row, column=7, value="")
        resumen.cell(row=summary_row, column=5).number_format = '"$" #,##0.0'
        for col in range(4, 8):
            resumen.cell(row=summary_row, column=col).font = Font(bold=True)
            resumen.cell(row=summary_row, column=col).alignment = Alignment(horizontal='center')
            resumen.cell(row=summary_row, column=col).border = thin_border

        # Column widths for D:G table
        resumen.column_dimensions['D'].width = 34
        resumen.column_dimensions['E'].width = 16
        resumen.column_dimensions['F'].width = 10
        resumen.column_dimensions['G'].width = 14

        # Print setup for summary sheet
        resumen.print_area = f"D2:G{summary_row}"
        resumen.page_setup.orientation = resumen.ORIENTATION_LANDSCAPE
        resumen.page_setup.paperSize = resumen.PAPERSIZE_A4
        resumen.sheet_properties.pageSetUpPr.fitToPage = True
        resumen.page_setup.fitToWidth = 1
        resumen.page_setup.fitToHeight = 0
        resumen.print_options.horizontalCentered = True
        resumen.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2)

        # Ensure continuous border around summary print block
        for r in range(2, summary_row + 1):
            for c in range(4, 8):
                cell = resumen.cell(row=r, column=c)
                if cell.border is None or cell.border == Border():
                    cell.border = thin_border

        from openpyxl.utils import get_column_letter

        for prof, rows in sorted(professor_nominas.items()):
            if not rows:
                continue
            
            df = pd.DataFrame(rows)
            sheet_name = prof[:31] if isinstance(prof, str) else str(prof)[:31]
            sheet_name = "".join(c for c in sheet_name if c not in '[]:*?/\\')
            ws = wb.create_sheet(title=sheet_name)
            
            # Row 1: Title "Nómina de Maestros YEAR"
            ws['A1'] = f"Nómina de Maestros {PAYROLL_YEAR}"
            ws['A1'].font = Font(bold=True, size=12)
            ws.merge_cells('A1:L1')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Row 2: Month/Year (parametrized)
            ws['A2'] = PAYROLL_MONTH
            ws['A2'].font = Font(bold=True, size=11)
            ws.merge_cells('A2:L2')
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Row 3: Professor name with instrument from dictionary
            instrument = PROFESSOR_INSTRUMENTS.get(prof, "Maestro")
            ws['A3'] = f"{instrument}: {prof}"
            ws['A3'].font = Font(bold=True, size=11)
            ws.merge_cells('A3:L3')
            ws['A3'].alignment = Alignment(horizontal='center')
            
            # Row 5: Column headers
            header_row = 5
            headers = ["No.", "G", "Alumno", "Día", "Horario", "Clase", "Duración", "Clases", "Pago Prop", "Monto", "Recibo"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', wrap_text=False)
            thin = Side(style="thin", color="000000")
            thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            # Data rows starting from row 7 (row 6 will be TOTAL after data loop)
            total_row = 4
            data_start_row = 6
            total_pago = 0
            total_monto = 0
            
            for row_idx, (_, row) in enumerate(df.iterrows(), start=data_start_row):
                classes_num = 0
                pago_num = 0
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Map DataFrame columns to header names
                    if header == "No.":
                        cell.value = row.get("No.", "")
                    elif header == "G":
                        cell.value = row.get("G", "")
                    elif header == "Alumno":
                        # Try various student name columns
                        cell.value = row.get("ALUMNO", row.get("Alumno", row.get("Completo", "")))
                    elif header == "Día":
                        cell.value = row.get("DÍA", row.get("Dia", ""))
                    elif header == "Horario":
                        horario_val = row.get("HORARIO", row.get("Horario", ""))
                        cell.value = _safe_time_value(horario_val)
                    elif header == "Clase":
                        cell.value = row.get("CLASE", row.get("Clase", ""))
                    elif header == "Duración":
                        cell.value = row.get("DURACIÓN", row.get("Duracion", ""))
                    elif header == "Clases":
                        cell.value = row.get("Clases", "")
                        try:
                            classes_num = float(cell.value) if cell.value not in (None, "", "0") else 0
                        except Exception:
                            classes_num = 0
                    elif header == "Pago Prop":
                        pago = row.get("Pago Prop", 0)
                        try:
                            pago_num = float(pago) if pago not in (None, "", "0") else 0
                        except Exception:
                            pago_num = 0
                        cell.value = pago_num
                        total_pago += pago_num
                        cell.number_format = '#,##0.0'
                    elif header == "Monto":
                        # Formula: Hx * Ix (Clases * Pago Prop)
                        cell.value = f"=H{row_idx}*I{row_idx}"
                        try:
                            total_monto += classes_num * pago_num
                        except Exception:
                            pass
                        cell.number_format = '#,##0.0'
                    elif header == "Recibo":
                        cell.value = row.get("RECIBO ACTUAL", row.get("Recibo", ""))
                    
                    cell.alignment = Alignment(horizontal='center', wrap_text=False)
                    cell.font = Font(size=9)  # Smaller font
                    cell.border = thin_border
            
            # Total row (at row 6, update with actual totals)
            last_data_row = data_start_row + len(df) - 1
            
            # Format TOTAL row: TOTAL | $ 720 | PESOS MXN
            # Column A: "TOTAL"
            ws.cell(row=total_row, column=9).value = "TOTAL"
            ws.cell(row=total_row, column=9).font = Font(bold=True, size=9)
            ws.cell(row=total_row, column=9).border = thin_border
            ws.cell(row=total_row, column=9).alignment = Alignment(horizontal='center', wrap_text=False)
            
            
            # Total Monto with currency symbol (column 10 = J)
            ws.cell(row=total_row, column=10).value = f"=SUM(J{data_start_row}:J{last_data_row})"
            ws.cell(row=total_row, column=10).font = Font(bold=True, size=9)
            ws.cell(row=total_row, column=10).number_format = '"$" #,##0.0'
            ws.cell(row=total_row, column=10).border = thin_border
            ws.cell(row=total_row, column=10).alignment = Alignment(horizontal='center', wrap_text=False)
            
            # Currency name (column 11 = K)
            ws.cell(row=total_row, column=11).value = CURRENCY_NAME
            ws.cell(row=total_row, column=11).font = Font(bold=True, size=9)
            ws.cell(row=total_row, column=11).border = thin_border
            ws.cell(row=total_row, column=11).alignment = Alignment(horizontal='center', wrap_text=False)
            
            # Footer section at bottom
            footer_start_row = last_data_row + 3
            
            # Professor name line
            ws.cell(row=footer_start_row, column=1).value = f"{instrument}: {prof}"
            ws.cell(row=footer_start_row, column=1).font = Font(bold=True)
            
            # Nómina month/year line (parametrized)
            ws.cell(row=footer_start_row + 1, column=1).value = f"Nómina {PAYROLL_MONTH} {PAYROLL_YEAR}"
            ws.cell(row=footer_start_row + 1, column=1).font = Font(bold=True)
            
            # Date on right side (5th of next month)
            month = PAYROLL_MONTH_NUM
            year = PAYROLL_YEAR
            next_month = month + 1 if month < 12 else 1
            next_year = year if month < 12 else year + 1
            date_str = f"5/{next_month}/{next_year}"
            ws.cell(row=footer_start_row, column=10).value = date_str
            ws.cell(row=footer_start_row, column=10).font = Font(bold=True)
            ws.cell(row=footer_start_row, column=10).alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 31
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 12
            ws.column_dimensions['L'].width = 12
            # Set print area: from A1 to one column after the "Recibo" column,
            # and down to the row containing the "Nómina {month} {year}" footer.
            try:
                try:
                    rec_index = headers.index("Recibo")
                except ValueError:
                    rec_index = len(headers) - 1
                # rec_index is 0-based; +1 => exact Recibo column as print boundary
                last_col_idx = rec_index + 1
                last_col = get_column_letter(last_col_idx)
                print_end_row = footer_start_row + 1
                ws.print_area = f"A1:{last_col}{print_end_row}"

                # Print setup for professor sheets
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                ws.print_options.horizontalCentered = True
                ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2)
            except Exception:
                pass
        
        wb.save(out_path)
        print(f"✓ Exported all nóminas into workbook: {out_path}")
    except Exception as e:
        print(f"✗ Error exporting single-workbook nómina: {e}")


def export_to_excel(df, filename):
    try:
        df.to_excel(filename, index=False)
        print(f"✓ Exported data to {filename}")
    except Exception as e:
        print(f"✗ Error exporting to Excel: {e}")


def export_to_csv(df, filename):
    try:
        df.to_csv(filename, index=False)
        print(f"✓ Exported data to {filename}")
    except Exception as e:
        print(f"✗ Error exporting to CSV: {e}")


def main():
    # Get the current script directory
    script_dir = Path(__file__).parent
    file_path = script_dir / EXCEL_FILE
    
    # Check if file exists
    if not file_path.exists():
        print(f"✗ Error: File '{EXCEL_FILE}' not found in {script_dir}")
        return

    # Load and filter data
    filtered_data = load_and_filter_data(file_path, SHEET_NAME, DATA_RANGE)


    
    if filtered_data is not None and not filtered_data.empty:
        # Define which columns you want to display (customize this list)
        # Example: showing key columns only
        columns_of_interest = [
            "ALUMNO",
            "PROFESOR",
            "CURSO",
            "DÍA",
            "HORARIO",
            "CLASE",
            "RECIBO ACTUAL"
        ]
        
        # Display the filtered data
        # Option 1: Show only specific columns
        # display_filtered_data(filtered_data, columns_of_interest)
        
        # Option 2: Show all columns (uncomment to use)
        # display_filtered_data(filtered_data)
        
        # GROUP STUDENTS BY PROFESSOR
        
        professor_groups = group_by_professor(filtered_data)
        
        print(f"\n✓ Found {len(professor_groups)} professors")
        print(f"  Professors: {', '.join(sorted(professor_groups.keys()))}\n")
        

        # Example: Access specific professor's students
        for professor_name in sorted(professor_groups.keys()):
            students_array = professor_groups[professor_name]
            print(f"✓ {professor_name}: {len(students_array)} students")
            # You can now work with students_array for each professor
            # students_array is a list of dictionaries, one per student


        professor_nominas = generate_payroll_for_professors(professor_groups)

        
        # Uncomment to export all data to one file:
        # export_to_excel(filtered_data, "estudiantes_activos.xlsx")
        # export_to_csv(filtered_data, "estudiantes_activos.csv")
        
        # Uncomment to export each professor to a separate file:
        # export_professors_to_separate_files(professor_groups, "professor_files")
        
        # Export all professors into a single workbook (one sheet per professor)
        xlsx_out = script_dir / f"{PAYROLL_MONTH}_Nomina.xlsx"
        export_nomina_single_workbook(professor_nominas, xlsx_out)
        ok_pdf, pdf_msg = export_xlsx_to_pdf(xlsx_out)
        if ok_pdf:
            print(f"✓ Exported PDF: {pdf_msg}")
        else:
            print(f"⚠ PDF export skipped: {pdf_msg}")
        # Summary statistics
        print(f"\n📊 Summary:")
        print(f"   Total active students: {len(filtered_data)}")
        print(f"   Total professors: {len(professor_groups)}")
        
        # Additional analysis examples
        if "CURSO" in filtered_data.columns:
            print(f"\n📚 Students by Course:")
            print(filtered_data["CURSO"].value_counts().to_string())
        
    
    else:
        print("No active students found or error loading data")


if __name__ == "__main__":
    # Launch the GUI by default for interactive browsing and generation
    def run_gui():
        # initialize customtkinter appearance
        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        root = ctk.CTk()
        root.title("Nómina Generator - GUI")

        frm = ctk.CTkFrame(root, corner_radius=6)
        frm.grid(row=0, column=0, padx=12, pady=12, sticky=(tk.N, tk.S, tk.E, tk.W))

        # File selection
        ctk.CTkLabel(frm, text="Excel file:").grid(row=0, column=0, sticky=tk.W)
        # Do not set a default path — user will choose file via Browse
        file_var = tk.StringVar(value="")
        ent_file = ctk.CTkEntry(frm, textvariable=file_var, width=600)
        ent_file.grid(row=0, column=1, sticky=(tk.W, tk.E))

        def browse_file():
            p = filedialog.askopenfilename(title="Select Excel file",
                                           filetypes=[("Excel files", "*.xls;*.xlsx;*.xlsm;*.xlsb"), ("All files", "*")])
            if p:
                file_var.set(p)

        ctk.CTkButton(frm, text="Browse...", command=browse_file).grid(row=0, column=2, sticky=tk.W, padx=(8,0))

        # Sheet name
        ctk.CTkLabel(frm, text="Sheet:").grid(row=1, column=0, sticky=tk.W)
        sheet_var = tk.StringVar(value=SHEET_NAME)
        ctk.CTkEntry(frm, textvariable=sheet_var).grid(row=1, column=1, sticky=(tk.W, tk.E))

        # Use range checkbox
        use_range_var = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(frm, text=f"Use data range ({DATA_RANGE})", variable=use_range_var).grid(row=2, column=1, sticky=tk.W)

        # Payroll month/year
        ctk.CTkLabel(frm, text="Payroll month:").grid(row=3, column=0, sticky=tk.W)
        month_var = tk.StringVar(value=PAYROLL_MONTH)
        ctk.CTkEntry(frm, textvariable=month_var, width=160).grid(row=3, column=1, sticky=tk.W)

        ctk.CTkLabel(frm, text="Year:").grid(row=3, column=2, sticky=tk.W)
        year_var = tk.IntVar(value=PAYROLL_YEAR)
        ctk.CTkEntry(frm, textvariable=year_var, width=80).grid(row=3, column=3, sticky=tk.W)

        # Log area (use tkinter ScrolledText for convenience)
        log = ScrolledText(frm, width=100, height=20)
        log.grid(row=6, column=0, columnspan=4, pady=(8, 0))

        def gui_log(msg):
            log.insert(tk.END, msg + "\n")
            log.see(tk.END)

        # Actions
        loaded_df = {"df": None}

        def do_load():
            path = Path(file_var.get())
            if not path.exists():
                messagebox.showerror("File not found", f"File not found: {path}")
                return
            sname = sheet_var.get() or None
            use_range = use_range_var.get()
            gui_log(f"Loading '{path}' sheet='{sname}' use_range={use_range}")
            df = load_and_filter_data(path, sname, DATA_RANGE if use_range else None)
            loaded_df['df'] = df
            if df is None:
                gui_log("Error loading data (see console)")
                return
            gui_log(f"Loaded {len(df)} rows")
            groups = group_by_professor(df)
            gui_log(f"Found {len(groups)} professors: {', '.join(sorted(groups.keys()))}")

        def do_generate():
            df = loaded_df.get('df')
            if df is None or df.empty:
                messagebox.showwarning("No data", "No loaded data. Please load a file first.")
                return
            groups = group_by_professor(df)
            gui_log("Generating nóminas...")
            prof_nom = generate_payroll_for_professors(groups)

            outdir = filedialog.asksaveasfilename(title="Save single workbook as...",
                                                  defaultextension='.xlsx',
                                                  initialfile=f"{month_var.get()}_Nomina.xlsx",
                                                  filetypes=[("Excel workbook", "*.xlsx")])
            if not outdir:
                gui_log("Export canceled")
                return
            try:
                export_nomina_single_workbook(prof_nom, outdir)
                ok_pdf, pdf_msg = export_xlsx_to_pdf(outdir)
                gui_log(f"Exported nóminas to {outdir}")
                if ok_pdf:
                    gui_log(f"Exported PDF to {pdf_msg}")
                    messagebox.showinfo("Exported", f"Exported nóminas to:\n{outdir}\n\nPDF:\n{pdf_msg}")
                else:
                    gui_log(f"PDF export skipped: {pdf_msg}")
                    messagebox.showwarning("Exported XLSX", f"Exported nóminas to:\n{outdir}\n\nPDF not exported:\n{pdf_msg}")
            except Exception as e:
                gui_log(f"Error exporting: {e}")
                messagebox.showerror("Error", str(e))

        def do_list_professors():
            df = loaded_df.get('df')
            if df is None:
                messagebox.showwarning("No data", "No loaded data. Please load a file first.")
                return
            groups = group_by_professor(df)
            gui_log("Professors and counts:")
            for p, students in sorted(groups.items()):
                gui_log(f" - {p}: {len(students)} students")
        # User must click Load & Filter after selecting a file
        ctk.CTkButton(frm, text="Load & Filter", command=do_load).grid(row=5, column=0, sticky=tk.W, pady=(8,0))
        ctk.CTkButton(frm, text="List Professors", command=do_list_professors).grid(row=5, column=1, sticky=tk.W, pady=(8,0))
        ctk.CTkButton(frm, text="Generate & Export", command=do_generate).grid(row=5, column=2, sticky=tk.W, pady=(8,0))

        # Make UI responsive
        root.grid_columnconfigure(0, weight=1)
        root.grid_rowconfigure(0, weight=1)
        frm.grid_columnconfigure(1, weight=1)

        root.mainloop()

    run_gui()
